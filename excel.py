import sys
from openpyxl import load_workbook

def main():
  # excel file load
  excel_path='/Users/taharayoshinobu/Documents/ProgramingStudy/python_excel/table2.xlsx'
  wb = load_workbook(filename=excel_path, read_only=True)

  f = open("insert.txt", "w")
    
  # displayed sheetname list
  # work_sheets
  ws = wb["SQL"]
  query = create_query(ws)
  for row in range(3, ws.max_row + 1):
    params = create_params(ws, row)
    # None→Null
    params = params.replace('None', 'NULL')
    sql = query + params
    print(sql)
    f.write(sql + "\n")

  # close excel file
  wb.close()
  f.close()
  

def create_query(ws):

  query = "insert into" + ws.cell(column=1, row=1).value + " ("

  for col in range(ws.min_column, ws.max_column + 1):
    query = query + ws.cell(column=col, row=2).value
    if col == ws.max_column:
      query = query + " )"
    else:
      query = query + " ,"
  return query

def create_params(ws, row):
  params = "values ("
  for col in range(ws.min_column, ws.max_column + 1):
    params = params + str(ws.cell(column=col, row=row).value)
    if col == ws.max_column:
      params = params + " )"
    else:
      params = params + " ,"

  return params
    
main()